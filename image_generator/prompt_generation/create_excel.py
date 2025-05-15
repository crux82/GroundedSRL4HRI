import pandas as pd
from PIL import Image
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment
import io

offset = 1
IMAGE_RESIZE_SHAPE = (400, 400) # (width, height)
SCALE_WIDTH_FACTOR = 1 / 7
SCALE_HEIGHT_FACTOR = 3 / 4

VERTICAL_AND_HORIZONTAL_CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)


def insert_images_to_excel(df, excel_file_path):
    # Create a new Excel workbook and sheet
    wb = Workbook()
    ws = wb.active

    column_names = [
        { 'Image ID' : 64 },
        { 'Annotated Image': 64 },
        { 'Prompt' : 170 },
        { 'Prompt Type' : 64 },
        { 'Huric ID' : 64 },
        { 'Important Entities To Include' : 130 },
        { 'Important Entities To Exclude' : 130 },
        { 'Optional Entities' : 128 },
        { 'Command' : 100 },
        { 'Frame Semantics' : 128 },
        { 'Spatial Relations' : 128 },
        { 'Score' : 64 },
        { 'Evalutation' : 120 }
    ]

    evaluation_column_index = -1
    for col_num, column_dict in enumerate(column_names, start=1):
        column_name, column_width = list(column_dict.items())[0]
        ws.cell(row=1, column=col_num).value = column_name
        ws.cell(row=1, column=col_num).alignment = VERTICAL_AND_HORIZONTAL_CENTER_ALIGNMENT
        if col_num == 2:
          ws.column_dimensions['B'].width = IMAGE_RESIZE_SHAPE[0] * SCALE_WIDTH_FACTOR
        else:
          ws.column_dimensions[chr(ord('A') + col_num - 1)].width = column_width * SCALE_WIDTH_FACTOR

        if column_name == 'Evalutation':
          evaluation_column_index = col_num
    

    list_of_output = [
      'immagine_inconsistente',
      'immagine_malformata',
      'entita_mancante',
      'entita_aggiunta',
      'entita_corrette_ma_stato_inconsistente',
      'OK'
    ]

    
    dv = DataValidation(type="list", formula1=f'"{",".join(list_of_output)}"', allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f'{chr(ord("A") + evaluation_column_index - 1)}1:{chr(ord("A") + evaluation_column_index - 1)}1048576')

    # Process each image path in the DataFrame
    for idx, row in df.iterrows():
        image_id = row['image_id']
        huric_id = row['huric_id']
        prompt = row['prompt']
        prompt_type = row['prompt_type']
        important_entities_to_include = row['important_entities_to_include']
        important_entities_to_exclude = row['important_entities_to_exclude']
        optional_entities = row['optional_entities']
        command = row['command']
        frame_semantics = row['frame_semantics']
        spatial_relations = row['spatial_relations']
        score = row['score']


        ws.cell(row=idx + offset + 1, column=1).value = image_id
        ws.cell(row=idx + offset + 1, column=3).value = prompt
        ws.cell(row=idx + offset + 1, column=4).value = prompt_type
        ws.cell(row=idx + offset + 1, column=5).value = huric_id
        ws.cell(row=idx + offset + 1, column=6).value = "\n".join(important_entities_to_include) if len(important_entities_to_include) > 0 else "None"
        ws.cell(row=idx + offset + 1, column=7).value = "\n".join(important_entities_to_exclude) if len(important_entities_to_exclude) > 0 else "None"
        ws.cell(row=idx + offset + 1, column=8).value = "\n".join(optional_entities) if len(optional_entities) > 0 else "None"
        ws.cell(row=idx + offset + 1, column=9).value = command
        ws.cell(row=idx + offset + 1, column=10).value = "\n".join(frame_semantics) if len(frame_semantics) > 0 else "None"
        ws.cell(row=idx + offset + 1, column=11).value = "\n".join(spatial_relations) if len(spatial_relations) > 0 else "None"
        ws.cell(row=idx + offset + 1, column=12).value = score


        image_path = row['labeled_image_path']  # Assuming the column name is 'image_path'

        # Open the image using PIL
        try:
            img = Image.open(image_path)
        except Exception as e:
            print(f"Error opening image {image_path}: {e}")
            continue
        
        # Resize image to IMAGE_RESIZE_SHAPE
        img_resized = img.resize(IMAGE_RESIZE_SHAPE)

        # Convert the PIL image to a format openpyxl can handle (like PNG in memory)
        img_byte_arr = io.BytesIO()
        img_resized.save(img_byte_arr, format='PNG')
        img_byte_arr.seek(0)

        # Create Openpyxl Image object
        openpyxl_img = OpenpyxlImage(img_byte_arr)

        # Insert the image into the worksheet, in the next free row
        cell = f'B{idx + offset + 1}'  # Row index starts from 1 in Excel, so we add 1
        openpyxl_img.anchor = cell
        ws.add_image(openpyxl_img, cell)
        ws.row_dimensions[offset + idx + 1].height = img_resized.height * SCALE_HEIGHT_FACTOR
        for col_num in range(1, len(column_names) + 1):
          ws.cell(row=idx + offset + 1, column=col_num).alignment = VERTICAL_AND_HORIZONTAL_CENTER_ALIGNMENT
        

    
    # Save the workbook to the specified path
    wb.save(excel_file_path)
    print(f"Excel file saved to {excel_file_path}")

# Example usage:
# Assuming you have a DataFrame 'df' with an 'image_path' column
data = {
  'image_id': [
      1,
      2,
      3
  ],
  'huric_id': [
      2175,
      3625,
      3555
  ],
  'command':[
      "please carry the mug to the bathroom",
      "grab the phone near the tv on the table",
      "take the phone on the bed in the living room"

  ],
  'prompt': [
      "A wide-angle shot of an apartment kitchen with a ceramic mug resting on a countertop near a stainless steel sink. The mug is positioned centrally in the frame, and in the background, a wooden cabinet and a hanging bag can be seen. The lighting is soft and natural, coming from an unseen window. The apartment has a warm and cozy aesthetic, with no visible people or robots.",
      "A wide shot of a modern apartment living room showcasing a television on a TV stand near a small coffee table. The coffee table holds a smartphone, positioned close to the television. A black recorder sits on the opposite side of the table, partially obscured by a decorative lamp. The walls are adorned with framed artwork, and a window with blinds partially open allows soft natural light to enter. The scene is devoid of people, robots, or dining tables.",
      "A wide-angle view of a modern apartment bedroom. The bed, covered with a blue comforter, is against a wall with a wooden headboard. A smartphone is resting on the bed near the center, placed close to a folded blanket. A closet door is slightly ajar in the background, and a poster hangs on the wall. A coffee cup is on a nightstand. The scene is well-lit by natural light from a window. No people or robots are in sight."
  ],
  'important_entities_to_include': [
      [],
      ["Television"],
      ["Bed", "Cellphone"],
  ],
  'important_entities_to_exclude': [
      ['Cup'],
      ["Cellphone", "Diningtable"],
      [],
  ],
  'optional_entities': [
      ["Plate", "Cabinet", "Pasta", "Can", "Bag"],
      ["Recorder", "Box", "Folder", "Plate"],
      ["Milk", "Coffee", "Coke", "Mayo", "Closet", "Poster"],
  ],
  'frame_semantics': [
      ["Frame: \"Bringing\" (evoked in the command by \"carry\")\n\t\t\u2022\tFrame Element: \"Theme\" \u2192 \"Cup\" (represented in the command by \"the mug\").\n\t\t\u2022\tFrame Element: \"Goal\" \u2192 \"Bathroom\" (represented in the command by \"to the bathroom\")."],
      ["Frame is **Taking** and the frame elements are:\n\t\t-Frame Element **Theme** with the semantic head **Cellphone**."],
      ["Frame: \"Bringing\" (evoked in the command by \"take\")\n\t\t\u2022\tFrame Element: \"Theme\" \u2192 \"Cellphone\" (represented in the command by \"the phone on the bed\").\n\t\t\u2022\tFrame Element: \"Goal\" \u2192 \"Room\" (represented in the command by \"in the living room\")."],
  ],
  'spatial_relations': [
      [],
      ["Television is close to Cellphone", "Television is close to Diningtable", "Cellphone is close to Television", "Cellphone is close to Diningtable", "Diningtable is close to Television", "Diningtable is close to Cellphone"],
      ["Bed is close to Cellphone", "Cellphone is close to Bed"],
  ],
  'bad': [0,0,0],
  'score': [0,0,0],
  'boxes': [
      [0,0,0,0],
      [0,0,0,0],
      [0,0,0,0]
  ],
  'image_path': [
      './images/2175_full_1_1_original.png',
      './images/3625_partial_1_1_original.png',
      './images/3555_full_1_1_original.png'
  ],
  'config_file': [
      './config/2175_full_1_1_config.json',
      './config/3625_partial_1_1_config.json',
      './config/3555_full_1_1_config.json'
  ],
  'prompt_type' : [
      'full',
      'partial',
      'full'
  ],
  'labeled_image_path': [
      './images/2175_full_1_1_annotated.png',
      './images/3625_partial_1_1_annotated.png',
      './images/3555_full_1_1_annotated.png'
  ]  # List your image paths here
}
df = pd.DataFrame(data)

# Specify the Excel file path
excel_file_path = 'images_in_excel.xlsx'

# Insert images into the Excel file
insert_images_to_excel(df, excel_file_path)
